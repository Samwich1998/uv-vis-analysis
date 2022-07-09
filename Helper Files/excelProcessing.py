#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Need to Install on the Anaconda Prompt:
    $ pip install pyexcel
"""


# Basic Modules
import os
import sys
# Read/Write to Excel
import csv
import pyexcel
import openpyxl as xl
# Openpyxl Styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font


class dataProcessing:        
        
    def xls2xlsx(self, excelFile, outputFolder):
        """
        Converts .xls Files to .xlsx Files That OpenPyxl Can Read
        If the File is Already a .xlsx Files, Do Nothing
        If the File is Neither a .xls Nor .xlsx, it Exits the Program
        """
        # Check That the Current Extension is .xls or .xlsx
        _, extension = os.path.splitext(excelFile)
        # If the Extension is .xlsx, the File is Ready; Do Nothing
        if extension == '.xlsx':
            return excelFile
        # If the Extension is Not .xls/.xlsx, Then the Data is in the Wrong Format; Exit Program
        if extension not in ['.xls', '.xlsx']:
            print("Cannot Convert File to .xlsx")
            sys.exit()
        
        # Create Output File Directory to Save Data ONLY If None Exists
        os.makedirs(outputFolder, exist_ok = True)
        # Convert '.xls' to '.xlsx'
        filename = os.path.basename(excelFile)
        newExcelFile = outputFolder + filename + "x"
        pyexcel.save_as(file_name = excelFile, dest_file_name = newExcelFile, logfile=open(os.devnull, 'w'))
        
        # Return New Excel File Name
        return newExcelFile
    
    def txt2csv(self, txtFile, csvFile, csvDelimiter = ",", overwriteCSV = False):
        # Check to See if CSV Conversion Alreayd Occurred
        if not os.path.isfile(csvFile) or overwriteCSV:
            with open(txtFile, "r") as inputData:
                in_reader = csv.reader(inputData, delimiter = csvDelimiter)
                with open(csvFile, 'w', newline='') as out_csv:
                    out_writer = csv.writer(out_csv)
                    for row in in_reader:
                        out_writer.writerow(row)
    
    def convertToExcel(self, inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = 0):
        """
        inputFile: The Input TXT/CSV File to Convert XLSX
        excelFile: The Output Excel File Name (XLSX)
        """
        # If the File is Not Already Converted: Convert the CSV to XLSX
        if not os.path.isfile(excelFile) or overwriteXL:
            # Make Excel WorkBook
            xlWorkbook = xl.Workbook()
            xlWorksheet = xlWorkbook.active
            # Write the Data from the CSV File to the Excel WorkBook
            with open(inputFile, "r") as inputData:
                inReader = csv.reader(inputData, delimiter = excelDelimiter)
                with open(excelFile, 'w+', newline=''):
                    for row in inReader:
                        xlWorksheet.append(row)
            # Save as New Excel File
            xlWorkbook.save(excelFile)
        # Else Load the Data from the Excel File
        else:
            # Load the Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum]
        
        # Return Excel Sheet
        return xlWorkbook, xlWorksheet
    

class saveData():
    
    def addExcelAesthetics(self, WB_worksheet):
        # Center the Data in the Cells
        align = Alignment(horizontal='center',vertical='center',wrap_text=True)        
        for column_cells in WB_worksheet.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            WB_worksheet.column_dimensions[xl.utils.get_column_letter(column_cells[0].column)].width = length
            
            for cell in column_cells:
                cell.alignment = align
        # Header Style
        for cell in WB_worksheet["1:1"]:
            cell.font = Font(color='00FF0000', italic=True, bold=True)
        
        return WB_worksheet
    
    def saveData(self, dataToSave, saveDataFolder, saveExcelName, sheetName = "UV-Vis Analysis"):
        print("Saving the Data")
        # Create Output File Directory to Save Data: If Not Already Created
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Create Path to Save the Excel File
        excelFile = saveDataFolder + saveExcelName
        
        # If the File is Not Present: Create it
        if not os.path.isfile(excelFile):
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        else:
            print("Excel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excelFile)
            WB_worksheet = WB.create_sheet(sheetName)
            
        # Add the Header
        headers = ["Sample Name", "Wavelength (nm)", "Absorbance (AU)", "Baseline Subtracted (AU)"]
        WB_worksheet.append(headers)
        
        # Organize and save the data
        for data in dataToSave:
            # Write the Data to Excel
            WB_worksheet.append(data)
        
        # Add Excel Aesthetics
        WB_worksheet = self.addExcelAesthetics(WB_worksheet)    
            
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()
    

class processFiles(dataProcessing):
    
    def getFiles(self, dataDirectory, fileDoesntContain, fileContains):
        # If Using All the CSV Files in the Folder
        dpvFiles = []; filesAdded = set();
        for fileName in os.listdir(dataDirectory):
            fileBase = os.path.splitext(fileName)[0]
            if fileName.endswith((".txt",'csv','xlsx', '.tsv')) and fileDoesntContain not in fileName and fileContains in fileName and fileBase not in filesAdded:
                dpvFiles.append(fileName)
                filesAdded.add(fileBase)
        if len(dpvFiles) == 0:
            print("No TXT/CSV/XLSX Files Found in the Data Folder:", dataDirectory)
            print("Found the Following Files:", os.listdir(dataDirectory))
            sys.exit()
        
        return dpvFiles

    def extractData_UVVis(self, uvVisWorksheet):
        
        newSample = False
        wavelengthList = [[]]; absorbanceList = [[]]; sampleNames = []
        # Loop Through the Info Section and Extract the Needxed Run Info from Excel
        rowGenerator = uvVisWorksheet.rows
        for cell in rowGenerator:
            # Get Cell Value
            cellVal = cell[0].value
            if cellVal == None:
                if newSample:
                    wavelengthList.append([])
                    absorbanceList.append([])
                    newSample = False
                continue
            
            # If there is a new sample ready
            if cellVal.startswith("Sample "):
                sampleNames.append(cellVal)
                newSample = True
            # If there is data to add
            elif newSample and cellVal.replace(".", "", 1).isdigit():
                wavelengthList[-1].append(float(cell[0].value))
                absorbanceList[-1].append(float(cell[1].value))
        
        if len(wavelengthList[-1]) == 0:
            wavelengthList.pop()
            absorbanceList.pop()
            
        return wavelengthList, absorbanceList, sampleNames
    
    
    def getData(self, oldFile, outputFolder, testSheetNum = 0, excelDelimiter = "\t"):
        """
        --------------------------------------------------------------------------
        Input Variable Definitions:
            oldFile: The Path to the Excel File Containing the Data: txt, csv, xls, xlsx
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(oldFile):
            print("The following Input File Does Not Exist:", oldFile)
            sys.exit()

        # Convert TXT and CSV Files to XLSX
        if oldFile.endswith((".txt", ".csv", '.tsv')):
            # Extract Filename Information
            oldFileExtension = os.path.basename(oldFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = outputFolder + "Excel Files/"
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok = True)

            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, xlWorksheet = self.convertToExcel(oldFile, excelFile, excelDelimiter = excelDelimiter, overwriteXL = True, testSheetNum = testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif oldFile.endswith(".xlsx"):
            excelFile = oldFile
            # Load the Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum]
        else:
            print("The Following File is Neither CSV, TXT, Nor XLSX:", excelFile)
            sys.exit()
        
        # Extract the Data
        print("Extracting Data from the Excel File:", excelFile)
        wavelengthList, absorbanceList, sampleNames = self.extractData_UVVis(xlWorksheet)
        
        xlWorkbook.close()
        # Finished Data Collection: Close Workbook and Return Data to User
        print("Done Collecting Data");
        return wavelengthList, absorbanceList, sampleNames


if __name__ == "__main__":
    
    inputFile = './diffusion_4.txt'
    processFiles().getData(inputFile)
    
    
    
    